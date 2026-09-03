"""Tests writer xlsx — structure du classeur produit à partir des fake
fixtures (3 jours anonymisés)."""
from __future__ import annotations

from pathlib import Path

import pytest
import yaml
from openpyxl import load_workbook

from balance_pdf_import.parser_planning import parser_pdf
from balance_pdf_import.writer_xlsx import (
    COLONNES_LISTE,
    COLONNES_PROPOSES,
    COLONNES_STAGIAIRES,
    ecrire_xlsx,
)


FIXTURES = Path(__file__).parent / "fixtures" / "s0"
CONFIG_PATH = Path(__file__).parent.parent / "config" / "fake-fixtures.yml"


@pytest.fixture
def xlsx_genere(tmp_path):
    with open(CONFIG_PATH) as f:
        config = yaml.safe_load(f)
    seances = []
    for pdf in sorted(FIXTURES.glob("*.pdf")):
        r = parser_pdf(pdf, config)
        seances.extend(r.seances)
    out = tmp_path / "balance.xlsx"
    ecrire_xlsx(seances, out)
    return out


def test_quatre_onglets_attendus(xlsx_genere):
    wb = load_workbook(xlsx_genere)
    assert wb.sheetnames == ["Liste", "Stagiaires", "Proposés", "Mode d'emploi"]


def test_onglet_liste_header_seul(xlsx_genere):
    """Liste = onglet des stagiaires (saisi par l'utilisateur). Le writer
    ne doit PAS peupler avec les morceaux du PDF (qui sont des morceaux
    d'intervenants) — ils vivent uniquement dans Proposés."""
    wb = load_workbook(xlsx_genere)
    ws = wb["Liste"]
    assert [c.value for c in ws[1]] == COLONNES_LISTE
    assert ws.max_row == 1


def test_onglet_stagiaires_header_seul(xlsx_genere):
    wb = load_workbook(xlsx_genere)
    ws = wb["Stagiaires"]
    assert [c.value for c in ws[1]] == COLONNES_STAGIAIRES
    assert ws.max_row == 1


def test_onglet_proposes_24_seances(xlsx_genere):
    wb = load_workbook(xlsx_genere)
    ws = wb["Proposés"]
    assert [c.value for c in ws[1]] == COLONNES_PROPOSES
    assert ws.max_row == 25  # header + 24 séances


def test_onglet_proposes_dates_iso(xlsx_genere):
    import re
    iso = re.compile(r"^\d{4}-\d{2}-\d{2}$")
    wb = load_workbook(xlsx_genere)
    ws = wb["Proposés"]
    for r in range(2, ws.max_row + 1):
        date = ws.cell(row=r, column=3).value
        assert iso.match(date), f"ligne {r} : date non-ISO {date!r}"


def test_onglet_proposes_salle_renseignee(xlsx_genere):
    """Le writer remplit la salle depuis le PDF — info précieuse pour le
    solveur, non-inventée."""
    wb = load_workbook(xlsx_genere)
    ws = wb["Proposés"]
    salles_uniques = set()
    for r in range(2, ws.max_row + 1):
        salle = ws.cell(row=r, column=6).value
        if salle:
            salles_uniques.add(salle)
    assert len(salles_uniques) >= 3  # au moins Le Pressoir, La Grange, Salle Nord
