"""Test end-to-end du CLI `balance-pdf-import` — de la ligne de commande
au xlsx sur disque + rapport d'audit, contre les fake fixtures."""
from __future__ import annotations

from pathlib import Path

from click.testing import CliRunner
from openpyxl import load_workbook

from balance_pdf_import.cli import main


FIXTURES = Path(__file__).parent / "fixtures" / "s0"
CONFIG_PATH = Path(__file__).parent.parent / "config" / "fake-fixtures.yml"


def test_cli_end_to_end_produit_xlsx_et_audit(tmp_path):
    out_xlsx = tmp_path / "balance.xlsx"
    out_audit = tmp_path / "balance.audit.txt"

    args = ["--config", str(CONFIG_PATH), "--out", str(out_xlsx), "--audit", str(out_audit)]
    for pdf in sorted(FIXTURES.glob("*.pdf")):
        args.extend(["--pdf", str(pdf)])

    result = CliRunner().invoke(main, args)
    assert result.exit_code == 0, f"CLI failed: {result.output}\n{result.exception}"
    assert out_xlsx.exists()
    assert out_audit.exists()

    wb = load_workbook(out_xlsx)
    assert wb["Proposés"].max_row == 25  # header + 24 séances
    assert wb["Liste"].max_row == 9      # header + 8 morceaux

    audit_txt = out_audit.read_text()
    assert "RAPPORT D'AUDIT" in audit_txt
    assert "Séances extraites          : 24" in audit_txt


def test_cli_option_pdf_dir(tmp_path):
    """--pdf-dir prend tous les *.pdf du dossier récursivement."""
    out_xlsx = tmp_path / "balance.xlsx"
    result = CliRunner().invoke(main, [
        "--config", str(CONFIG_PATH),
        "--pdf-dir", str(FIXTURES),
        "--out", str(out_xlsx),
    ])
    assert result.exit_code == 0, f"CLI failed: {result.output}\n{result.exception}"
    assert out_xlsx.exists()

    wb = load_workbook(out_xlsx)
    assert wb["Proposés"].max_row == 25


def test_cli_aucun_pdf_erreur(tmp_path):
    """Ni --pdf ni --pdf-dir → erreur explicite."""
    result = CliRunner().invoke(main, [
        "--config", str(CONFIG_PATH),
        "--out", str(tmp_path / "out.xlsx"),
    ])
    assert result.exit_code != 0
    assert "Aucun PDF" in result.output


def test_cli_signale_pdf_manquant(tmp_path):
    """Chemin PDF inexistant → click exit code non-zéro."""
    out_xlsx = tmp_path / "balance.xlsx"
    result = CliRunner().invoke(main, [
        "--config", str(CONFIG_PATH),
        "--pdf", "/tmp/does-not-exist.pdf",
        "--out", str(out_xlsx),
    ])
    assert result.exit_code != 0


# ── A1 audit Stéphane 2026-08-30 ─────────────────────────────────────────

def test_cli_audit_contient_liste_fichiers_traites(tmp_path):
    """A1 : le rapport doit contenir la liste explicite des PDFs
    réellement ouverts avec pages + séances par fichier — permet à l'user
    de vérifier que ce qu'il a fourni a bien été lu."""
    out_xlsx = tmp_path / "balance.xlsx"
    out_audit = tmp_path / "balance.audit.txt"
    args = ["--config", str(CONFIG_PATH), "--out", str(out_xlsx), "--audit", str(out_audit)]
    for pdf in sorted(FIXTURES.glob("*.pdf")):
        args.extend(["--pdf", str(pdf)])
    result = CliRunner().invoke(main, args)
    assert result.exit_code == 0
    audit_txt = out_audit.read_text()
    assert "Fichiers PDF traités" in audit_txt
    # Chaque fichier fixture doit apparaître nommément
    for pdf in FIXTURES.glob("*.pdf"):
        assert pdf.name in audit_txt


def test_cli_audit_deterministe_meme_hash_2_exec(tmp_path):
    """A2 : 2 exécutions identiques → 2 rapports au même hachage.
    Permet de comparer avant/après un fix, essentiel pour validation."""
    import hashlib

    def hachage(chemin: Path) -> str:
        return hashlib.sha256(chemin.read_bytes()).hexdigest()

    def executer(out_dir: Path) -> Path:
        out_xlsx = out_dir / "balance.xlsx"
        out_audit = out_dir / "balance.audit.txt"
        args = ["--config", str(CONFIG_PATH), "--out", str(out_xlsx), "--audit", str(out_audit)]
        for pdf in sorted(FIXTURES.glob("*.pdf")):
            args.extend(["--pdf", str(pdf)])
        result = CliRunner().invoke(main, args)
        assert result.exit_code == 0
        return out_audit

    (tmp_path / "run1").mkdir()
    (tmp_path / "run2").mkdir()
    a1 = executer(tmp_path / "run1")
    a2 = executer(tmp_path / "run2")
    assert hachage(a1) == hachage(a2), "audit non déterministe entre 2 exécutions identiques"


def test_cli_refuse_xlsx_si_zero_seance(tmp_path):
    """A1 : si config incompatible avec PDFs → 0 séance → refuse produire
    xlsx (échec bruyant vaut mieux qu'un résultat plausible et faux)."""
    # Config qui ne matche AUCUNE salle des fake fixtures.
    config_bidon = tmp_path / "config-bidon.yml"
    config_bidon.write_text(
        "session: Test\n"
        "dates: {dimanche: 2026-04-12}\n"
        "salles: ['Salle Inexistante A', 'Salle Inexistante B']\n"
        "mots_cles_seance: ['Répétition']\n"
        "ignorer: []\n"
        "motif_responsable: 'avec (?P<nom>.+?)(?:\\s|$)'\n"
    )
    out_xlsx = tmp_path / "balance.xlsx"
    out_audit = tmp_path / "balance.audit.txt"
    args = ["--config", str(config_bidon), "--out", str(out_xlsx), "--audit", str(out_audit)]
    for pdf in sorted(FIXTURES.glob("*.pdf")):
        args.extend(["--pdf", str(pdf)])
    result = CliRunner().invoke(main, args)
    assert result.exit_code == 2, f"expected exit 2, got {result.exit_code}: {result.output}"
    assert not out_xlsx.exists(), "xlsx ne doit pas être produit si 0 séance"
    # L'audit doit être écrit tout de même — permet diagnostic sans re-run
    assert out_audit.exists()
    assert "Fichiers PDF traités" in out_audit.read_text()


def test_cli_autorise_zero_seance_avec_flag(tmp_path):
    """A1 : --autoriser-zero-seance permet de produire un xlsx vide
    quand le user le souhaite explicitement (mode dev/debug)."""
    config_bidon = tmp_path / "config-bidon.yml"
    config_bidon.write_text(
        "session: Test\n"
        "salles: ['Salle X']\n"
        "mots_cles_seance: ['Répétition']\n"
        "ignorer: []\n"
        "motif_responsable: 'avec (?P<nom>.+?)(?:\\s|$)'\n"
    )
    out_xlsx = tmp_path / "balance.xlsx"
    args = [
        "--config", str(config_bidon),
        "--out", str(out_xlsx),
        "--autoriser-zero-seance",
    ]
    for pdf in sorted(FIXTURES.glob("*.pdf")):
        args.extend(["--pdf", str(pdf)])
    result = CliRunner().invoke(main, args)
    assert result.exit_code == 0
    assert out_xlsx.exists()


def test_cli_bandeau_mode_demo_dans_audit(tmp_path):
    """A1 : si config = fake-fixtures.yml → bandeau MODE DÉMONSTRATION
    en tête de rapport, visible dès la 1ère ligne."""
    out_xlsx = tmp_path / "balance.xlsx"
    out_audit = tmp_path / "balance.audit.txt"
    args = ["--config", str(CONFIG_PATH), "--out", str(out_xlsx), "--audit", str(out_audit)]
    for pdf in sorted(FIXTURES.glob("*.pdf")):
        args.extend(["--pdf", str(pdf)])
    result = CliRunner().invoke(main, args)
    assert result.exit_code == 0
    audit_txt = out_audit.read_text()
    # Le bandeau démo est en tout premier (avant même le titre RAPPORT).
    assert "MODE DÉMONSTRATION" in audit_txt.split("═")[0]


def test_cli_compteur_seances_sans_fin_correct(tmp_path):
    """A4 : le compteur « séances sans heure de fin » doit inclure les
    séances écartées (fin manquante) — ne pas afficher 0 quand le détail
    contient des lignes fin manquante."""
    out_xlsx = tmp_path / "balance.xlsx"
    out_audit = tmp_path / "balance.audit.txt"
    args = ["--config", str(CONFIG_PATH), "--out", str(out_xlsx), "--audit", str(out_audit)]
    for pdf in sorted(FIXTURES.glob("*.pdf")):
        args.extend(["--pdf", str(pdf)])
    result = CliRunner().invoke(main, args)
    assert result.exit_code == 0
    audit_txt = out_audit.read_text()
    # Le fake corpus a une répétition lundi 22:00 sans fin — le compteur
    # doit valoir au moins 1.
    lignes_compteur = [l for l in audit_txt.split("\n") if "Séances sans heure de fin" in l]
    assert lignes_compteur, "ligne compteur manquante"
    ligne = lignes_compteur[0]
    # Format « Séances sans heure de fin  : 1 (0 extraites, 1 écartées) »
    assert "0 " not in ligne.split(":")[1].strip()[:3], \
        f"compteur affiche 0 alors qu'une séance a la fin manquante: {ligne}"


def test_cli_comptage_par_morceau_dans_audit(tmp_path):
    """A5 (contrôle méthode) : le rapport contient le comptage par
    morceau pour permettre à l'user de vérifier N séances attendues
    vs extraites."""
    out_xlsx = tmp_path / "balance.xlsx"
    out_audit = tmp_path / "balance.audit.txt"
    args = ["--config", str(CONFIG_PATH), "--out", str(out_xlsx), "--audit", str(out_audit)]
    for pdf in sorted(FIXTURES.glob("*.pdf")):
        args.extend(["--pdf", str(pdf)])
    result = CliRunner().invoke(main, args)
    assert result.exit_code == 0
    audit_txt = out_audit.read_text()
    assert "Comptage séances par morceau" in audit_txt


def test_est_config_demo_detecte_fake_fixtures():
    """A1 : est_config_demo() reconnaît fake-fixtures.yml."""
    from balance_pdf_import.cli import est_config_demo
    assert est_config_demo(Path("/some/path/fake-fixtures.yml")) is True
    assert est_config_demo(Path("/some/path/vraie-config.yml")) is False
    assert est_config_demo(Path("/some/path/s5-2026.yml")) is False
