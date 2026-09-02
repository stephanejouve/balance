"""Writer xlsx — produit le classeur d'import Balance à partir des séances
extraites du PDF.

Contrat de sortie (brief) : onglets `Liste`, `Stagiaires`, `Proposés`,
`Mode d'emploi`, en-têtes strictement identiques au template. Rien
n'est inventé — les colonnes non renseignables depuis le PDF (pupitres,
tonalités, indispos) restent vides.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from .parser_planning import Seance


COLONNES_LISTE = [
    "Morceau", "Auteur", "Style", "Tona", "Resp", "Cherche",
    "Chant", "Piano", "Basse", "Batterie", "Guitare", "Vents",
]
COLONNES_STAGIAIRES = [
    "Nom", "Pupitre", "Pupitres additionnels",
    "Instrument", "Latéralité", "Indispos",
]
COLONNES_PROPOSES = ["Morceau", "Membres", "Date", "Début", "Fin", "Salle"]

# Mode d'emploi recopié de app/src/io/excel-template.ts pour respecter la
# source de vérité unique du template Balance (bug de dérive à éviter).
MODE_EMPLOI = [
    ("Exemple", "Que ça signifie", True),
    ("— Onglet Liste : cellules d'un pupitre —", None, True),
    ("(cellule vide)", "Pas de personne à ce pupitre pour ce morceau (défaut)", False),
    ("NON", "Explicitement pas ce pupitre (équivalent à vide, plus lisible)", False),
    ("CHERCHE", "Poste à pourvoir — les renforts compatibles seront suggérés dans l'UI", False),
    ("Emma, Bianca", "Plusieurs personnes au même pupitre (séparateur virgule)", False),
    ("Colette (contrebasse)", "Parenthèse = instrument : precision libre appliquée à cette personne", False),
    ("Pierre (SIG), Pierre (L)", "Parenthèses discriminantes : deux personnes distinctes (homonymes du même prénom)", False),
    (None, None, False),
    ("— Onglet Stagiaires : colonne Indispos —", None, True),
    ("mercredi 09h-10h chant", "Jour de la semaine + plage horaire + rôle ciblé", False),
    ("mardi 14:30 - 16:00", "Jour et plage horaire seuls", False),
    ("convalescence", "Texte libre conservé sans horaire spécifique (motif pour relecture humaine)", False),
    (None, None, False),
    ("— Onglet Stagiaires : intervenants inclus —", None, True),
    (
        "Bertrand au piano d'un morceau de stagiaires",
        "L'onglet Stagiaires accueille toutes les personnes du stage, intervenants compris. "
        "Un intervenant qui vient jouer dans un morceau se déclare ici comme les autres, avec son pupitre.",
        False,
    ),
    (None, None, False),
    ("— Onglet Proposés : concert du vendredi —", None, True),
    ("1 ligne = 1 séance", "Plusieurs séances du même morceau : autant de lignes avec le même titre", False),
    ("Date : 2026-08-28 ou 28/08/2026", "ISO ou format FR — les deux sont tolérés à l'import", False),
    ("Heure : 09:00 ou 9h30", "HH:MM ou format libre avec « h »", False),
    (None, None, False),
    ("— Exemple concret d'une ligne Liste —", None, True),
    ("Morceau", "Love", False),
    ("Auteur", "Nat King Cole", False),
    ("Style", "Jazz", False),
    ("Resp", "Emma", False),
    ("Chant", "Emma (B), Bianca (B)", False),
    ("Piano", "Prune", False),
    ("Basse", "Rose", False),
    ("Batterie", "CHERCHE", False),
    ("Guitare", "Cyril", False),
    ("Vents", "Serge", False),
]


def ecrire_xlsx(seances: list[Seance], chemin_sortie: Path) -> None:
    """Écrit le classeur Balance dans `chemin_sortie` à partir des séances."""
    wb = Workbook()
    wb.remove(wb.active)
    gras = Font(bold=True)

    ws_liste = wb.create_sheet("Liste")
    for i, col in enumerate(COLONNES_LISTE, start=1):
        cell = ws_liste.cell(row=1, column=i, value=col)
        cell.font = gras
    # Onglet Liste laissé vide : le PDF n'apporte que les morceaux
    # d'intervenants (grille du stage) qui vivent dans Proposés. Écrire
    # ces morceaux ici créait des Groupe fantômes que le solveur essayait
    # de placer en double des Impose. Liste est destinée aux stagiaires,
    # que l'utilisateur y saisit après import.

    ws_stag = wb.create_sheet("Stagiaires")
    for i, col in enumerate(COLONNES_STAGIAIRES, start=1):
        cell = ws_stag.cell(row=1, column=i, value=col)
        cell.font = gras

    ws_prop = wb.create_sheet("Proposés")
    for i, col in enumerate(COLONNES_PROPOSES, start=1):
        cell = ws_prop.cell(row=1, column=i, value=col)
        cell.font = gras
    seances_triees = sorted(seances, key=lambda s: (s.date, s.debut, s.morceau))
    for r, s in enumerate(seances_triees, start=2):
        ws_prop.cell(row=r, column=1, value=s.morceau)
        ws_prop.cell(row=r, column=3, value=s.date)
        ws_prop.cell(row=r, column=4, value=s.debut)
        if s.fin:
            ws_prop.cell(row=r, column=5, value=s.fin)
        ws_prop.cell(row=r, column=6, value=s.salle)

    ws_me = wb.create_sheet("Mode d'emploi")
    for r, (col1, col2, is_bold) in enumerate(MODE_EMPLOI, start=1):
        c1 = ws_me.cell(row=r, column=1, value=col1)
        c2 = ws_me.cell(row=r, column=2, value=col2)
        if is_bold:
            c1.font = gras
            c2.font = gras

    wb.save(chemin_sortie)
